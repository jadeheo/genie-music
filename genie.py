import requests
from bs4 import BeautifulSoup

from openpyxl import load_workbook

work_book = load_workbook('genie.xlsx')
work_sheet = work_book['music']

from pymongo import MongoClient
client = MongoClient('localhost', 27017)
db = client.dbsparta

headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = requests.get('https://www.genie.co.kr/chart/top200?ditc=D&rtm=N&ymd=20190908', headers=headers)

soup = BeautifulSoup(data.text, 'html.parser')

music_list = soup.select('.list-wrap > tbody > tr')

work_sheet.cell(row=1, column=1, value='순위')
work_sheet.cell(row=1, column=2, value='곡 제목')
work_sheet.cell(row=1, column=3, value='가수')


row = 2

for music in music_list:
    list_tag = music.select_one('td.info > a')
    if list_tag != None:
        ranking = music.select_one('td.number')
        unwanted = (ranking.find('span')).extract()
        #unwanted.extract()

        title = list_tag.text
        singer = music.select_one('td.info > a.artist.ellipsis').text
        #print(ranking.text, title, singer)

        db.music.insert_one({'ranking': ranking.text, 'title': title, 'singer': singer})

        work_sheet.cell(row=row, column=1, value=ranking.text)
        work_sheet.cell(row=row, column=2, value=title)
        work_sheet.cell(row=row, column=3, value=singer)

        row += 1


work_book.save('genie.xlsx')